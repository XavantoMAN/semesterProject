import datetime
import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import openpyxl


def get_today_currencies():
    today_currencies = pd.read_csv('today_currencies.csv')
    return today_currencies


def get_gdp_in_monetary_current():
    gdp_in_monetary = pd.read_csv('gdp_in_monetary_current.csv')
    return gdp_in_monetary


def get_key_rate_and_inflation():
    key_rate = pd.read_csv('key_rate_and_inflation.csv')
    return key_rate


def get_gdp_all_time():
    workbook = openpyxl.load_workbook('gdp_all_time.xlsx')
    sheetnames = workbook.sheetnames
    worksheet = workbook[sheetnames[1]]
    lst1 = []
    lst2 = []
    for i in range(1, 18):
        for j in range(3, 5):
            if j == 3:
                lst1.append(worksheet.cell(row=j, column=i).value)
            else:
                lst2.append(worksheet.cell(row=j, column=i).value)
    worksheet = workbook[sheetnames[2]]
    for i in range(1, 14):
        for j in range(4, 6):
            if j == 4:
                data = worksheet.cell(row=j, column=i).value
                if i > 11:
                    data = data[:-2]
                lst1.append(data)
            else:
                lst2.append(worksheet.cell(row=j, column=i).value)
    gdp_all_time = pd.DataFrame({'Year': lst1,
                                 'Amount': lst2})
    return gdp_all_time


def get_gdp_growth():
    gdp_all_time = get_gdp_all_time()
    lst1 = []
    iterator = 0
    prev = 1
    for i in gdp_all_time.Amount:
        if iterator == 0:
            lst1.append(0)
        else:
            curr = i
            value = round(((curr / prev) - 1) * 100, 2)
            lst1.append(value)
        prev = i
        iterator += 1
    gdp_growth_all_time = pd.DataFrame({'Year': gdp_all_time.Year,
                                        'Percent': lst1})
    return gdp_growth_all_time


def parse_gdp_all_time():
    url = 'https://rosstat.gov.ru/storage/mediabank/VVP_god_s_1995-2023.xlsx'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'
    }
    response = requests.get(url, headers=headers)
    print(response)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    with open('gdp_all_time.xlsx', 'wb') as f:
        f.write(response.content)


def parse_today_currencies():
    url = 'https://cbr.ru/currency_base/daily/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'
    }
    response = requests.get(url, headers=headers)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    print(response)
    parsed_html = bs(response.text, 'html.parser')
    table = parsed_html.find('table', 'data')
    table_head_temp = table.find_all('th')
    table_head = []
    table_data_temp = table.find_all('td')
    table_data = []
    currency = []
    currency_rate = []
    count = []
    letter_code = []
    num_code = []
    iterator = 1
    for i in table_head_temp:
        table_head.append(i.text)
    for i in table_data_temp:
        table_data.append(i.text)
    for i in range(len(table_data)):
        if iterator == 5:
            temp = float(table_data[i].replace(',', '.'))
            if count[-1] != 1:
                temp = temp/count[-1]
            currency_rate.append(temp)
            iterator = 0
        elif iterator == 4:
            currency.append(table_data[i])
        elif iterator == 3:
            count.append(int(table_data[i]))
        elif iterator == 2:
            letter_code.append(table_data[i])
        elif iterator == 1:
            num_code.append(table_data[i])
        iterator += 1
    temp_dict = {table_head[0]: num_code,
                 table_head[1]: letter_code,
                 table_head[2]: count,
                 table_head[3]: currency,
                 table_head[4]: currency_rate}
    today_currencies_table = pd.DataFrame(temp_dict)
    today_currencies_table.to_csv('today_currencies.csv', index=False)


def parse_key_rate_and_inflation():
    url = 'https://cbr.ru/hd_base/infl/?UniDbQuery.Posted=True&UniDbQuery.From=17.09.2013&UniDbQuery.To=31.05.2024'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'
    }
    response = requests.get(url, headers=headers)
    print(response)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    parsed_html = bs(response.text, 'html.parser')
    table = parsed_html.find('table', 'data')
    table_head_temp = table.find_all('th')
    table_head = []
    table_data_temp = table.find_all('td')
    table_data = []
    iterator = 1
    lst1 = []
    lst2 = []
    lst3 = []
    lst4 = []
    for i in table_head_temp:
        table_head.append(i.text)
    for i in table_data_temp:
        table_data.append(i.text)
    for i in table_data:
        if iterator == 1:
            date = i.split('.')
            lst1.append(datetime.date(int(date[1]), int(date[0]), 1))
        elif iterator == 2:
            lst2.append(float(i.replace(',', '.')))
        elif iterator == 3:
            lst3.append(float(i.replace(',', '.')))
        elif iterator == 4:
            if i != ' — ':
                lst4.append(float(i.replace(',', '.')))
            else:
                lst4.append(None)
            iterator = 0
        iterator += 1
    temp_dict = {table_head[0]: lst1,
                 table_head[1]: lst2,
                 table_head[2]: lst3,
                 table_head[3]: lst4}
    key_rate_and_inflation = pd.DataFrame(temp_dict)
    key_rate_and_inflation.to_csv('key_rate_and_inflation.csv', index=False)


def parse_gdp_in_monetary_current():
    url = 'https://ru.tradingeconomics.com/russia/gdp'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'}
    response = requests.get(url, headers=headers)
    print(response)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    parsed_html = bs(response.text, 'html.parser')
    table = parsed_html.find('table', 'table table-hover')
    table_data_temp = table.find_all('td')
    table_data = []
    lst1 = []
    lst2 = []
    lst3 = []
    lst4 = []
    iterator = 1
    for i in table_data_temp:
        table_data.append(i.text.strip(' \n\r'))
    for i in table_data:
        if iterator == 1:
            lst1.append(i)
        elif iterator == 2:
            lst2.append(float(i))
        elif iterator == 3:
            lst3.append(i)
        elif iterator == 4:
            lst4.append(i)
            iterator = 0
        iterator += 1
    temp_dict = {
        'Name': lst1,
        'Last': lst2,
        'Unit': lst3,
        'Date': lst4
    }
    gdp_in_monetary_current = pd.DataFrame(temp_dict)
    gdp_in_monetary_current.to_csv('gdp_in_monetary_current.csv', index=False)


def update_data():
    parse_today_currencies()
    parse_key_rate_and_inflation()
    parse_gdp_in_monetary_current()


def parse_currencies_and_codes():
    url = 'https://cbr.ru/currency_base/dynamics/'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'}
    response = requests.get(url, headers=headers)
    print(response)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    parsed_html = bs(response.text, 'html.parser')
    table = parsed_html.find('label', 'input_label')
    data = table.select('option[value]')
    with open('currency_names.csv', 'w') as f:
        for i in data:
            f.write(','+i.getText().strip())
    with open('currency_codes.csv', 'w') as f:
        for i in data:
            f.write(','+i.get('value'))


def parse_currency_for_period(currency_code: str, from_date: str, to_date: str):
    url = ('https://cbr.ru/currency_base/dynamics/?UniDbQuery.Posted=True&UniDbQuery.so=1&UniDbQuery.'
           f'mode=1&UniDbQuery.date_req1=&UniDbQuery.date_req2=&UniDbQuery.VAL_NM_RQ={currency_code}&UniDbQuery.'
           f'From={from_date}&UniDbQuery.To={to_date}')
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPad; CPU OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148'}
    response = requests.get(url, headers=headers)
    print(response)
    while response.status_code != 200:
        response = requests.get(url, headers=headers)
    parsed_html = bs(response.text, 'html.parser')
    table = parsed_html.find('table', 'data')
    table_data = table.find_all('td')
    lst1 = []
    lst2 = []
    lst3 = []
    iterator = 1
    for i in range(1, len(table_data)):
        if table_data[i].text.__contains__('\n'):
            continue
        if iterator == 1:
            lst1.append(table_data[i].text)
        elif iterator == 2:
            lst2.append(table_data[i].text)
        elif iterator == 3:
            iterator = 0
            temp = table_data[i].text.replace(',', '.')
            temp = temp.replace(' ', '')
            temp = round(float(temp), 2)
            lst3.append(temp/int(lst2[-1]))
        iterator += 1
    currency_for_period = pd.DataFrame({'Дата': lst1,
                                        'Единиц': lst2,
                                        'Курс': lst3})
    currency_for_period = currency_for_period[::-1]
    return currency_for_period
